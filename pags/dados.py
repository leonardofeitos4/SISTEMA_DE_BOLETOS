import openpyxl
import pandas as pd


arquivo_excel = 'boletos_TESTES.xlsx'

def get_data() -> pd.DataFrame:
    with pd.ExcelFile('boletos_TESTES.xlsx') as xls:
        return pd.read_excel(xls)

def get_clientes():
    df = get_data()
    
    return list(df.NOME_CLIENTE.unique())

def get_infos_cliente(cliente):
    df = get_data()
    
    mercados = list(df.query(f'NOME_CLIENTE == "{cliente}"').NOME_MERCADO.unique())
    
    telefones = list(df.query(f'NOME_CLIENTE == "{cliente}"').TELEFONE.unique())
    
    return mercados, telefones

def get_datas(cliente):
    df= get_data()
    
    datas = list(df.query(f'NOME_CLIENTE == "{cliente}"').DATA_VENCIMENTO.unique())
    
    return datas

def get_valores(cliente):
    df= get_data()
    
    valores = list(df.query(f'NOME_CLIENTE == "{cliente}"').VALOR.unique())
    
    return valores

def get_situacao(cliente, mercado, valor, data):
    df = get_data()
    
    situacao = df.query(f'NOME_CLIENTE == "{cliente}" and NOME_MERCADO == "{mercado}" and VALOR == {valor} and DATA_VENCIMENTO == "{data}"').SITUACAO.to_list()[0]
    print(situacao)
    
    return situacao


def adicionar_linha_excel(dados):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active
    ultima_linha = planilha.max_row + 1

    for index, item in enumerate(dados):
        planilha.cell(row=ultima_linha, column=index+1, value=item)
    

    workbook.save(arquivo_excel)


def deletar_linha_excel(nome_cliente, nome_mercado, valor,data_vencimento):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active

    
    linha_a_deletar = None
    for row in range(2, planilha.max_row + 1):
        if (
            planilha.cell(row=row, column=1).value == nome_cliente and
            planilha.cell(row=row, column=2).value == nome_mercado and
            planilha.cell(row=row, column=3).value == data_vencimento and
            planilha.cell(row=row, column=5).value == valor
        ):
            linha_a_deletar = row
            break

    if linha_a_deletar is not None:
        planilha.delete_rows(linha_a_deletar, 1)

        workbook.save(arquivo_excel)
        return True
    return False


def editar_linha_excel(dados_antigo, novos):
    workbook = openpyxl.load_workbook(arquivo_excel)
    planilha = workbook.active
    
    linha_a_editar = None
    for row in range(2, planilha.max_row + 1):
        if (
            planilha.cell(row=row, column=1).value == dados_antigo[0] and
            planilha.cell(row=row, column=2).value == dados_antigo[1] and
            planilha.cell(row=row, column=3).value == dados_antigo[2] and
            planilha.cell(row=row, column=5).value == dados_antigo[3]
        ):
            for colum, novo in novos:
                planilha.cell(row=row, column=colum, value=novo)
            workbook.save(arquivo_excel)
            return True
    return False
    